import streamlit as st
import pandas as pd
import asyncio
import re
from io import BytesIO
from datetime import datetime

import google.generativeai as genai
from openai import AsyncOpenAI

# =========================
# CONFIG
# =========================
st.set_page_config(page_title="Sistema IA Manutenção", layout="wide")

FICHEIRO_HIST = "historico.csv"

# API KEYS
GEMINI_API = st.secrets["GEMINI_API"]
OPENROUTER_API = st.secrets["OPENROUTER_API"]

# INIT APIs
genai.configure(api_key=GEMINI_API)

client_or = AsyncOpenAI(
    base_url="https://openrouter.ai/api/v1",
    api_key=OPENROUTER_API,
)

# =========================
# FUNÇÕES AUXILIARES
# =========================

def buscar_valor(campo, texto):
    match = re.search(rf"{campo}:\s*(.*)", texto)
    return match.group(1).strip() if match else "N/A"


def extrair_confianca(texto):
    if not texto:
        return 0
    try:
        return int(re.sub(r"\D", "", buscar_valor("CONFIANÇA", texto)))
    except:
        return 0


def criar_excel(texto, confianca):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    ws['A1'] = "RELATÓRIO AUTOMÁTICO"
    ws['A3'] = buscar_valor("ID", texto)
    ws['A4'] = buscar_valor("DATA", texto)
    ws['A5'] = buscar_valor("RESPONSÁVEL", texto)
    ws['A6'] = buscar_valor("EQUIPAMENTO", texto)
    ws['A7'] = buscar_valor("MODELO", texto)
    ws['A9'] = buscar_valor("DIAGNÓSTICO", texto)
    ws['A12'] = buscar_valor("SOLUÇÃO", texto)
    ws['A15'] = buscar_valor("GRAVIDADE", texto)
    ws['A16'] = f"{confianca}%"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def guardar_historico(dados):
    try:
        df = pd.read_csv(FICHEIRO_HIST)
    except:
        df = pd.DataFrame()

    df = pd.concat([df, pd.DataFrame([dados])])
    df.to_csv(FICHEIRO_HIST, index=False)


def carregar_historico():
    try:
        return pd.read_csv(FICHEIRO_HIST)
    except:
        return pd.DataFrame()

# =========================
# IA
# =========================

async def chamar_openrouter(prompt):
    try:
        res = await client_or.chat.completions.create(
            model="openrouter/free",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=600,
        )
        return res.choices[0].message.content
    except:
        return None


async def chamar_gemini(prompt):
    try:
        model = genai.GenerativeModel("gemini-1.5-flash")
        res = model.generate_content(prompt)
        return res.text
    except:
        return None


async def analisar(conteudo):
    prompt = f"""
Tu és um Engenheiro de Manutenção.

Analisa o relatório:

{conteudo}

Responde EXCLUSIVAMENTE neste formato:

--------------------------------------------------
ID:
DATA:
RESPONSÁVEL:
EQUIPAMENTO:
MODELO:
DIAGNÓSTICO:
SOLUÇÃO:
GRAVIDADE:
CONFIANÇA:
--------------------------------------------------
"""

    tarefa1 = chamar_gemini(prompt)
    tarefa2 = chamar_openrouter(prompt)
    tarefa3 = chamar_openrouter(prompt)

    r1, r2, r3 = await asyncio.gather(tarefa1, tarefa2, tarefa3)

    respostas = [r1, r2, r3]

    melhor = None
    melhor_conf = 0

    for r in respostas:
        conf = extrair_confianca(r)
        if conf > melhor_conf:
            melhor = r
            melhor_conf = conf

    return melhor, melhor_conf, r1, r2, r3


# =========================
# UI
# =========================

st.title("🏭 Sistema Inteligente de Manutenção")

menu = st.sidebar.selectbox("Menu", [
    "📤 Processar Relatórios",
    "📊 Dashboard",
    "📁 Histórico"
])

# =========================
# PROCESSAR
# =========================

if menu == "📤 Processar Relatórios":

    ficheiros = st.file_uploader(
        "Carregar relatórios",
        accept_multiple_files=True
    )

    if ficheiros:
        for ficheiro in ficheiros:

            st.subheader(f"📄 {ficheiro.name}")

            conteudo = ficheiro.read().decode("utf-8", errors="ignore")

            with st.spinner("A analisar com IA..."):
                resultado, conf, r1, r2, r3 = asyncio.run(analisar(conteudo))

            if resultado:
                st.code(resultado)
                st.success(f"Confiança: {conf}%")

                with st.expander("Ver respostas das IAs"):
                    st.code(r1)
                    st.code(r2)
                    st.code(r3)

                # Guardar histórico
                guardar_historico({
                    "ficheiro": ficheiro.name,
                    "data": datetime.now(),
                    "equipamento": buscar_valor("EQUIPAMENTO", resultado),
                    "gravidade": buscar_valor("GRAVIDADE", resultado),
                    "confianca": conf
                })

                # Excel
                excel = criar_excel(resultado, conf)

                st.download_button(
                    label="⬇️ Download Excel",
                    data=excel,
                    file_name=f"{ficheiro.name}_relatorio.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            else:
                st.error("Erro na análise")

# =========================
# DASHBOARD
# =========================

elif menu == "📊 Dashboard":

    df = carregar_historico()

    if df.empty:
        st.warning("Sem dados")
    else:
        col1, col2, col3 = st.columns(3)

        col1.metric("Total relatórios", len(df))
        col2.metric("Equipamentos únicos", df["equipamento"].nunique())

        criticos = len(df[df["gravidade"] == "Crítica"])
        col3.metric("Críticos", criticos)

        st.subheader("Gravidade")
        st.bar_chart(df["gravidade"].value_counts())

        st.subheader("Equipamentos")
        st.bar_chart(df["equipamento"].value_counts())

# =========================
# HISTÓRICO
# =========================

elif menu == "📁 Histórico":

    df = carregar_historico()

    if df.empty:
        st.warning("Sem histórico")
    else:

        filtro = st.selectbox(
            "Filtrar equipamento",
            ["Todos"] + list(df["equipamento"].unique())
        )

        if filtro != "Todos":
            df = df[df["equipamento"] == filtro]

        st.dataframe(df)
